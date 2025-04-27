from flask import Flask, render_template, request, url_for
import os
import openpyxl
import warnings
from datetime import datetime

# openpyxlの警告を非表示にする
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

app = Flask(__name__)

# Excelファイルの置き場所
EXCEL_DIR = './data'
RESULTS_PER_PAGE = 20  # 1ページあたりの表示件数

# 日付のデータから時刻を削除する関数
def format_date(value):
    if isinstance(value, datetime):
        return value.date()  # 時刻を削除して日付のみ返す
    return value

# 複数キーワードでAND検索する関数
def contains_all_keywords(row, keywords):
    for keyword in keywords:
        found = False
        for cell in row:
            if cell and keyword.lower() in str(cell).lower():
                found = True
                break
        if not found:
            return False
    return True

# 指定ディレクトリ内のExcelファイルを検索（.xlsxと.xlsmに対応）
def search_excel_files(keywords):
    results = []
    for filename in os.listdir(EXCEL_DIR):
        if filename.endswith(('.xlsx', '.xlsm')):
            filepath = os.path.join(EXCEL_DIR, filename)
            wb = openpyxl.load_workbook(filepath, data_only=True)
            # シート名が「メイン画面」のシートを対象
            if 'メイン画面' in wb.sheetnames:
                sheet = wb['メイン画面']
                for row in sheet.iter_rows(values_only=True):
                    # AND条件で検索
                    if contains_all_keywords(row, keywords):
                        # C列〜K列（0-indexed: 2〜10）を抽出し、日付を整形
                        row_data = [format_date(cell) for cell in row[2:11]]
                        results.append({
                            'values': row_data
                        })
            wb.close()
    return list(reversed(results))  # 結果を逆順にする

@app.route('/', methods=['GET', 'POST'])
def index():
    results = []
    keyword = ''
    page = int(request.args.get('page', 1))  # 現在のページ番号（デフォルトは1）
    
    if request.method == 'POST':
        keyword = request.form['keyword']
        keywords = [k.strip() for k in keyword.split() if k.strip()]  # キーワードをスペースで区切ってリスト化
        results = search_excel_files(keywords)
    else:
        # キーワードが入力されている場合、結果をフィルタリング
        if 'keyword' in request.args:
            keyword = request.args['keyword']
            keywords = [k.strip() for k in keyword.split() if k.strip()]
            results = search_excel_files(keywords)
    
    # 総検索結果行数を計算
    total_results = len(results)

    # ページごとの結果を取得
    start = (page - 1) * RESULTS_PER_PAGE
    end = start + RESULTS_PER_PAGE
    paginated_results = results[start:end]

    # 次ページ、前ページへのリンクを作成
    next_page = url_for('index', page=page + 1, keyword=keyword) if end < total_results else None
    prev_page = url_for('index', page=page - 1, keyword=keyword) if page > 1 else None

    return render_template('index.html', keyword=keyword, results=paginated_results, 
                           total_results=total_results, next_page=next_page, prev_page=prev_page, page=page)

if __name__ == '__main__':
    if not os.path.exists('templates'):
        os.makedirs('templates')
    with open('templates/index.html', 'w', encoding='utf-8') as f:
        f.write("""
<!DOCTYPE html>
<html>
<head>
    <meta charset="UTF-8">
    <title>Excelファイル検索</title>
    <style>
        body { font-family: sans-serif; margin: 2em; }
        input[type=text] { width: 300px; padding: 5px; }
        table { border-collapse: collapse; margin-top: 1em; }
        th, td { border: 1px solid #aaa; padding: 5px; }
        .pagination { margin-top: 20px; }
        .pagination a { margin: 0 5px; text-decoration: none; }
    </style>
</head>
<body>
    <h1>Excelファイル検索</h1>
    <form method="POST">
        <input type="text" name="keyword" placeholder="キーワードを入力" value="{{ keyword }}">
        <button type="submit">検索</button>
    </form>

    {% if results %}
        <h2>検索結果（総件数: {{ total_results }}件）</h2>
        <table>
            <tr>
                <th>C列〜K列の値</th>
            </tr>
            {% for item in results %}
                <tr>
                    <td>
                        {% for val in item['values'] %}
                            {{ val if val is not none else '' }}{% if not loop.last %} / {% endif %}
                        {% endfor %}
                    </td>
                </tr>
            {% endfor %}
        </table>
    {% elif keyword %}
        <p>一致するデータが見つかりませんでした。</p>
    {% endif %}

    <div class="pagination">
        {% if prev_page %}
            <a href="{{ prev_page }}">前のページ</a>
        {% endif %}
        {% if next_page %}
            <a href="{{ next_page }}">次のページ</a>
        {% endif %}
    </div>
</body>
</html>
""")
    app.run(debug=True)
